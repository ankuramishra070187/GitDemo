import openpyxl


# ******** give path of file ************************
workbook = openpyxl.load_workbook("/Users/ami23772/PycharmProjects/PythonSelFramework/TestData/PythonDemo.xlsx")
# ********* which sheet to be used ******************
sheet = workbook.active
# ********* selecting cells *************************
cell = sheet.cell(row=1, column=2)
# ******** to fetch value of the object present in that cell *****
print(cell.value)
print(sheet['C1'].value)
sheet.cell(row=2, column=2).value = "Ankur"
print(sheet.cell(row=2, column=2).value)
# ****** to know maximal filled rows and columns *********
print(sheet.max_row, sheet.max_column)
first_row = {}
second_row = {}

# ***** for loop to print all values in excel sheet **********
"""
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i,
                      column=1).value == "Testcase2":  # this will be used when i need specific row to be printed and ignore all other
            for j in range(2, sheet.max_column + 1):  # to get columns
                print(sheet.cell(row=i, column=j).value)
            # j += 1
        # i += 1
"""

# ******* Adding values to dictionary ***********************

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        # this will be used when i need specific row to be printed and ignore all other
        for j in range(2, sheet.max_column + 1):  # to get columns
            second_row[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    elif sheet.cell(row=i, column=1).value == "Testcase1":
        for j in range(2, sheet.max_column + 1):  # to get columns
            first_row[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(first_row, "\n", second_row)
# return first_row, second_row
