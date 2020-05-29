# from TestData.excelDemo import excelDemo
import openpyxl


class HomePageData():
    test_HomePage_Data = [{"firstname": "Ankur", "email": "mishra@ankur.com", "password": "sfdsffsf", "gender": "Male",
                           "birthdate": "07-01-1965"},
                          {"firstname": "Gabriella", "email": "singh@ankur.com", "password": "sfdsffsf",
                           "gender": "Female", "birthdate": "01-07-1967"}]

    # test_HomePage_Data = [excelDemo.createDictionary().first_row, excelDemo.createDictionary().second_row]

    @staticmethod
    def getTestData(test_case_name1, test_case_name2):
        first_row = {}
        second_row = {}
        workbook = openpyxl.load_workbook("/Users/ami23772/PycharmProjects/PythonSelFramework/TestData/PythonDemo.xlsx")
        # ********* which sheet to be used ******************
        sheet = workbook.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name2:
                # this will be used when i need specific row to be printed and ignore all other
                for j in range(2, sheet.max_column + 1):  # to get columns
                    second_row[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            elif sheet.cell(row=i, column=1).value == test_case_name1:
                for j in range(2, sheet.max_column + 1):  # to get columns
                    first_row[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [first_row, second_row]
